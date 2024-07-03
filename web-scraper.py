from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import re
import openpyxl

def find_and_click_load_more_button(driver):
    """
    Searches for the "Load More Results" button and clicks it if found.

    Args:
        driver (selenium.webdriver.Chrome): The Selenium WebDriver instance.

    Returns:
        bool: True if the button was found and clicked, False otherwise.
    """

    try:
        load_more_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//a[text()='LOAD MORE RESULTS']"))
        )
        load_more_button.click()
        return True
    except:
        return False

options = Options()
options.add_experimental_option('detach', True)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get('https://www.findaphotographer.com/search/results/eyJyZXN1bHRfcGFnZSI6InNlYXJjaFwvcmVzdWx0cyIsImRpc3RhbmNlOnRvIjoibGF0aXR1ZGV8bG9uZ2l0dWRlIiwiZGlzdGFuY2U6dW5pdCI6Im1pIiwibG9uZ2l0dWRlIjoibm90IElTX0VNUFRZIiwiY2F0ZWdvcnkiOiIzfDI4fDMwfDE0fDEwfDIwfDI2fDQ0fDIyfDMxfDMyfDEyfDMzfDIzfDM1fDM2fDE3fDM4fDQwfDE4IiwiZGlzdGFuY2U6cmFkaXVzIjoiMjUiLCJvcmRlcmJ5X3NvcnQiOiJmaXJzdF9uYW1lfGFzYyJ9')
driver.maximize_window()

num_attempts = 0  # Track the number of attempts to find the button
max_attempts = 5  # Set a maximum number of attempts to avoid infinite loops

while num_attempts < max_attempts:
    # Wait for results to load before checking for the button

    if not find_and_click_load_more_button(driver):
        # Button not found, break the loop
        break
    else:
        num_attempts = 0

      # print(num_attempts)
    num_attempts += 1

# Extract desired HTML content (replace with your specific logic)
html_content = driver.page_source  # Example: get entire page source
soup = BeautifulSoup(html_content, 'lxml') # Parse the HTML to create the soup

# Identify the element containing the info you want
data_elements = soup.find_all('div', class_="search_result")

# Prepare data for .xlsx file
class Photographer:
    def __init__(self, name, address, link):
        self.name = name
        self.address = address
        self.link = link

    def __str__(self):
      return f"Name: {self.name}, Address: {self.address}, Link: {self.link}\n"

data = []
for element in data_elements:
    # Extract specific data from each element
    name = element.find('strong').text.strip()
    address_element = element.find('br').next_sibling
    address = re.sub(r"\s+", " ", address_element.text.strip()) 

    if len(address) == 0:
        br_tags = element.find_all('br')
        if len(br_tags) >= 2:
            second_br = br_tags[1]
            address = second_br.next_sibling
            address = re.sub(r"\s+", " ", address.text.strip())     

    anchor_tag = element.find('strong').parent
    link = "https://www.findaphotographer.com" + anchor_tag.get('href')

    photographer = Photographer(name, address, link)
    data.append(photographer)



# for photographer in data:
#     print (photographer)

# Store the data in an excel spreadsheet
wb = openpyxl.Workbook()
ws = wb.active # Get the active worksheet

# Set column headers
ws.cell(row=1, column=1).value = 'Name'
ws.cell(row=1, column=2).value = 'Address'
ws.cell(row=1, column=3).value = 'Link'

# Write data starting from row 2
row_number = 2
for photographer in data:
    ws.cell(row=row_number, column=1).value = photographer.name
    ws.cell(row=row_number, column=2).value = photographer.address
    ws.cell(row=row_number, column=3).value = photographer.link
    row_number += 1


# Save the workbook
wb.save('photographers.xlsx')
print("Excel spreadsheet created successfully")


driver.quit()